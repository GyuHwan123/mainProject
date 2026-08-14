import json
from io import BytesIO

import httpx
from fastapi import APIRouter, Depends, File, Form, HTTPException, Query, UploadFile
from fastapi.responses import Response, StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from pydantic import BaseModel, Field

from app.api.routes.auth import require_current_user
from app.core.config import settings
from app.models.user import User
from app.schemas.ocr import DocumentHistoryItem, OCRResponse
from app.services.supabase_service import supabase_service
from app.services.pii_service import privacy_boxes

router = APIRouter()
MAX_FILE_SIZE = 50 * 1024 * 1024


class WorkbookExportRequest(BaseModel):
    title: str = Field(default="추출 문서", max_length=120)
    rows: list[list[str]] = Field(min_length=1, max_length=1000)


def save_result(*, user: User, file: UploadFile, content: bytes, result: OCRResponse, upload_origin: str = "OCR") -> OCRResponse:
    document = supabase_service.save_ocr_document(
        user_email=user.email,
        filename=file.filename or result.filename,
        mime_type=file.content_type or "application/octet-stream",
        content=content,
        content_type=result.content_type,
        pages=[page.model_dump() for page in result.pages],
        upload_origin=upload_origin,
    )
    result.document_id = document["id"]
    return result


@router.post("/archive", response_model=OCRResponse)
async def archive_extracted_file(
    file: UploadFile = File(...),
    result_json: str = Form(...),
    user: User = Depends(require_current_user),
) -> OCRResponse:
    try:
        result = OCRResponse.model_validate(json.loads(result_json))
    except (json.JSONDecodeError, ValueError) as exc:
        raise HTTPException(status_code=400, detail="잘못된 추출 결과입니다.") from exc
    content = await file.read()
    if len(content) > MAX_FILE_SIZE:
        raise HTTPException(status_code=413, detail="파일은 최대 50MB까지 저장할 수 있습니다.")
    return save_result(user=user, file=file, content=content, result=result)


@router.post("/upload", response_model=OCRResponse)
async def upload_file(
    file: UploadFile = File(...),
    ground_truth_json: str | None = Form(default=None),
    upload_origin: str = Query(default="OCR", pattern="^(OCR|RAG)$"),
    processing_mode: str = Query(default="document", pattern="^(document|receipt)$"),
    user: User = Depends(require_current_user),
) -> OCRResponse:
    content = await file.read()
    if len(content) > MAX_FILE_SIZE:
        raise HTTPException(status_code=413, detail="파일은 최대 50MB까지 저장할 수 있습니다.")

    filename = file.filename or "upload"
    mime_type = file.content_type or "application/octet-stream"
    try:
        async with httpx.AsyncClient(timeout=300.0) as client:
            response = await client.post(
                f"{settings.OCR_BASE_URL.rstrip('/')}/upload",
                params={"processing_mode": processing_mode},
                files={"file": (filename, content, mime_type)},
                data={"ground_truth_json": ground_truth_json} if ground_truth_json else None,
            )
            response.raise_for_status()
    except httpx.HTTPStatusError as exc:
        raise HTTPException(status_code=502, detail=exc.response.text or "OCR processing failed") from exc
    except httpx.RequestError as exc:
        raise HTTPException(status_code=503, detail="OCR service is unavailable") from exc

    result = OCRResponse.model_validate(response.json())
    return save_result(user=user, file=file, content=content, result=result, upload_origin=upload_origin)


@router.post("/docx-preview")
async def preview_docx(
    file: UploadFile = File(...),
    _user: User = Depends(require_current_user),
) -> Response:
    content = await file.read()
    if len(content) > MAX_FILE_SIZE:
        raise HTTPException(status_code=413, detail="파일은 최대 50MB까지 업로드할 수 있습니다.")
    if not (file.filename or "").lower().endswith(".docx"):
        raise HTTPException(status_code=400, detail="DOCX 파일만 미리보기할 수 있습니다.")

    try:
        async with httpx.AsyncClient(timeout=120.0) as client:
            preview = await client.post(
                f"{settings.OCR_BASE_URL.rstrip('/')}/docx-preview",
                files={
                    "file": (
                        file.filename or "document.docx",
                        content,
                        file.content_type or "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
                    )
                },
            )
            preview.raise_for_status()
    except httpx.HTTPStatusError as exc:
        raise HTTPException(status_code=502, detail=exc.response.text or "DOCX preview failed") from exc
    except httpx.RequestError as exc:
        raise HTTPException(status_code=503, detail="OCR service is unavailable") from exc

    return Response(content=preview.content, media_type="application/pdf")


@router.post("/spreadsheet-preview", response_model=OCRResponse)
async def preview_spreadsheet(
    file: UploadFile = File(...),
    _user: User = Depends(require_current_user),
) -> OCRResponse:
    content = await file.read()
    if len(content) > MAX_FILE_SIZE:
        raise HTTPException(status_code=413, detail="파일은 최대 50MB까지 업로드할 수 있습니다.")
    if not (file.filename or "").lower().endswith((".xlsx", ".xlsm")):
        raise HTTPException(status_code=400, detail="XLSX 또는 XLSM 파일만 미리보기할 수 있습니다.")

    try:
        async with httpx.AsyncClient(timeout=120.0) as client:
            preview = await client.post(
                f"{settings.OCR_BASE_URL.rstrip('/')}/spreadsheet-preview",
                files={"file": (file.filename or "spreadsheet.xlsx", content, file.content_type)},
            )
            preview.raise_for_status()
    except httpx.HTTPStatusError as exc:
        raise HTTPException(status_code=502, detail=exc.response.text or "Excel preview failed") from exc
    except httpx.RequestError as exc:
        raise HTTPException(status_code=503, detail="OCR service is unavailable") from exc

    return OCRResponse.model_validate(preview.json())


@router.get("/history", response_model=list[DocumentHistoryItem])
def ocr_history(
    upload_origin: str | None = Query(default=None, pattern="^(OCR|RAG)$"),
    user: User = Depends(require_current_user),
) -> list[DocumentHistoryItem]:
    return [DocumentHistoryItem.model_validate(item) for item in supabase_service.list_ocr_documents(user.email, upload_origin)]


@router.get("/documents/{document_id}", response_model=OCRResponse)
def get_document(document_id: str, user: User = Depends(require_current_user)) -> OCRResponse:
    document = supabase_service.get_ocr_document(user.email, document_id)
    pages = document.get("bounding_boxes") or []
    return OCRResponse(
        document_id=document["id"],
        filename=document["file_name"],
        content_type="stored_document",
        pages=pages,
    )


@router.get("/documents/{document_id}/file")
def get_document_file(document_id: str, user: User = Depends(require_current_user)) -> StreamingResponse:
    document = supabase_service.get_ocr_document(user.email, document_id)
    content, mime_type = supabase_service.download_document(document["file_url"])
    return StreamingResponse(BytesIO(content), media_type=mime_type)


@router.get("/documents/{document_id}/privacy-boxes")
def get_privacy_boxes(document_id: str, user: User = Depends(require_current_user)) -> list[dict]:
    document = supabase_service.get_ocr_document(user.email, document_id)
    return privacy_boxes(document.get("bounding_boxes") or [])


@router.post("/export-workbook")
def export_workbook(
    payload: WorkbookExportRequest,
    _user: User = Depends(require_current_user),
) -> StreamingResponse:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = (payload.title.strip() or "추출 문서")[:31]
    column_count = max((len(row) for row in payload.rows), default=1)
    header_fill = PatternFill("solid", fgColor="1F4E78")
    alternate_fill = PatternFill("solid", fgColor="EAF2F8")
    thin = Side(style="thin", color="B8C4D0")
    table_border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def excel_value(value: str):
        text = str(value or "").strip()
        numeric = text.replace(",", "").replace("₩", "").replace("원", "").strip()
        if numeric and numeric.lstrip("+-").replace(".", "", 1).isdigit():
            return float(numeric) if "." in numeric else int(numeric)
        return text

    for row in payload.rows:
        worksheet.append([excel_value(value) for value in row] + [""] * (column_count - len(row)))
    for row_index, row in enumerate(worksheet.iter_rows(min_row=1, max_row=len(payload.rows), max_col=column_count), start=1):
        for cell in row:
            cell.border = table_border
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            if isinstance(cell.value, (int, float)):
                cell.number_format = '#,##0.00;[Red]-#,##0.00'
            if row_index == 1:
                cell.fill = header_fill
                cell.font = Font(color="FFFFFF", bold=True)
            elif row_index % 2 == 0:
                cell.fill = alternate_fill
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    for column in worksheet.columns:
        letter = column[0].column_letter
        worksheet.column_dimensions[letter].width = min(
            max((len(str(cell.value or "")) for cell in column), default=8) + 2,
            48,
        )
    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="extracted-document.xlsx"'},
    )
