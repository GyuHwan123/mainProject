from __future__ import annotations

import logging
import re
from difflib import SequenceMatcher
from io import BytesIO
from time import perf_counter
from typing import Any

from openpyxl import load_workbook

from app.api.routes.finance import _classify_receipt_with_model, _normalize, _normalize_expense_category
from app.services.finance_error_analysis_service import analyze_finance_evaluation_failure
from app.services.finance_normalization import normalization_equivalent
from app.services.finance_workbook_service import HEADERS_BY_TYPE, SHEET_NAMES, SUMMARY_SHEET_NAME, build_finance_workbook


logger = logging.getLogger(__name__)
CORE_FIELDS = (
    "document_type", "expense_category", "merchant", "transaction_date",
    "supply_amount", "tax_amount", "total_amount", "payment_method",
    "total_quantity", "discount_amount", "card_number",
)
NUMBER_FIELDS = {
    "supply_amount", "tax_amount", "total_amount", "quantity", "unit_price",
    "total_quantity", "discount_amount",
}

# Explicit receipt-domain aliases keep evaluation deterministic while allowing
# equivalent item names written in different languages. Keys are compacted by
# removing spaces and punctuation before lookup; values are stable concepts.
ITEM_NAME_ALIASES = {
    "haircut": "beauty_service",
    "haircutservice": "beauty_service",
    "hairsalon": "beauty_service",
    "beautysalon": "beauty_service",
    "hairdressingservice": "beauty_service",
    "미용서비스": "beauty_service",
    "헤어컷": "beauty_service",
    "커트": "beauty_service",
    "미용실": "beauty_service",
    "barbershop": "barber_service",
    "barberservice": "barber_service",
    "이발서비스": "barber_service",
    "이발소": "barber_service",
    # Taxi receipts use several labels for the same purchased transport
    # service. Keep these exact aliases narrow so unrelated bus/train fares do
    # not become equivalent merely because they also contain ``요금``.
    "탑승요금": "taxi_transport_service",
    "택시이용": "taxi_transport_service",
    "택시요금": "taxi_transport_service",
    "택시운임": "taxi_transport_service",
    "택시승차요금": "taxi_transport_service",
}

ITEM_TOKEN_ALIASES = {
    "brushed": "브러쉬드",
    "alpaca": "알파카",
    "peru": "페루",
}
ITEM_IGNORED_TOKENS = {
    "diy", "도안", "상품", "제품",
    "best", "추천", "강추", "초강추", "오리지널",
}
MERCHANT_IGNORED_DESCRIPTORS = ("중고서점",)

def _normalize_item_display_name(value: Any) -> str:
    """Remove receipt-only decorations without correcting OCR spelling."""
    text = " ".join(str(value or "").strip().split())
    text = re.sub(r"^\s*\d{1,3}\s*[).:\-]\s*", "", text)
    promotional_prefix = re.compile(
        r"^\s*\+?\s*(?:(?:초강추|강추|추천|best)\s*[:：\-]?\s*)+",
        re.IGNORECASE,
    )
    had_promotional_prefix = promotional_prefix.search(text) is not None
    text = promotional_prefix.sub("", text)
    if had_promotional_prefix:
        text = re.sub(r"^\s*오리지널\s+", "", text, flags=re.IGNORECASE)
    # Product identifiers and retail barcodes are evidence metadata, not part
    # of the human-readable product name used by the ground truth.
    text = re.sub(r"\s+(?:[A-Z]{1,5}\d{6,}|\d{8,})\s*$", "", text, flags=re.IGNORECASE)
    return " ".join(text.split())


def _item_tokens(value: Any) -> set[str]:
    text = _normalize_item_display_name(value).lower()
    for source, target in ITEM_TOKEN_ALIASES.items():
        text = re.sub(rf"\b{re.escape(source)}\b", f" {target} ", text)
    return {
        token for token in re.findall(r"[0-9]+(?:[.,][0-9]+)?|[a-z]+|[가-힣]+", text)
        if token not in ITEM_IGNORED_TOKENS
    }


def normalize_ground_truth(truth: dict[str, Any]) -> dict[str, Any]:
    """Convert Korean receipt labels to the English schema used by the model.

    Classification labels are deliberately not inferred from ``카테고리``.  An
    explicitly supplied English ``document_type``/``expense_category`` remains
    supported for existing evaluation payloads.
    """
    normalized = {
        field: truth[field]
        for field in CORE_FIELDS
        if field in truth
    }

    korean_to_english = {
        "가게명": "merchant",
        "총 물품 수량": "total_quantity",
        "할인액": "discount_amount",
        "총 결제액": "total_amount",
        "카테고리": "expense_category",
        "결제방식": "payment_method",
        "카드번호": "card_number",
    }
    for korean_key, english_key in korean_to_english.items():
        if english_key not in normalized and korean_key in truth:
            normalized[english_key] = truth[korean_key]

    if "expense_category" in normalized:
        normalized["expense_category"] = _normalize_expense_category(normalized["expense_category"])

    # 할인액은 선택 필드다. 정답에서 생략된 경우 모델도 값을 만들지
    # 않았는지 평가할 수 있도록 명시적인 null로 정규화한다.
    if "discount_amount" not in normalized:
        normalized["discount_amount"] = None

    raw_date = truth.get("transaction_date", truth.get("구매일자"))
    if raw_date is not None:
        date_text = str(raw_date).strip()
        normalized["transaction_date"] = date_text[:10] if date_text else None

    source_items = truth.get("items")
    if not isinstance(source_items, list):
        source_items = truth.get("구매물품")
    if isinstance(source_items, list):
        normalized["items"] = []
        for source_item in source_items:
            if not isinstance(source_item, dict):
                continue
            item = {}
            aliases = {
                "name": "상품명",
                "quantity": "수량",
                "unit_price": "단가",
                "total_amount": "금액",
            }
            for english_key, korean_key in aliases.items():
                if english_key in source_item:
                    item[english_key] = source_item[english_key]
                elif korean_key in source_item:
                    item[english_key] = source_item[korean_key]
            normalized["items"].append(item)

    return normalized


def _canonical(field: str, value: Any) -> Any:
    if field in NUMBER_FIELDS:
        try:
            cleaned = re.sub(r"[^0-9.+-]", "", str(value or 0).replace(",", ""))
            return round(float(cleaned or 0), 2)
        except (TypeError, ValueError):
            return 0.0
    if value is None:
        return ""
    text = " ".join(str(value).strip().lower().split())
    if field == "name":
        text = _normalize_item_display_name(text)
        compact = re.sub(r"[^0-9a-z가-힣]", "", text)
        return ITEM_NAME_ALIASES.get(compact, text)
    if field == "transaction_date":
        parts = re.findall(r"\d+", text)[:3]
        if len(parts) == 3:
            return f"{int(parts[0]):04d}-{int(parts[1]):02d}-{int(parts[2]):02d}"
    if field == "payment_method":
        compact = re.sub(r"[^0-9a-z가-힣]", "", text)
        aliases = {
            "cash": ("cash", "현금", "현금결제"),
            "credit_card": ("creditcard", "card", "카드", "신용카드", "카드결제", "법인카드"),
            "debit_card": ("debitcard", "checkcard", "체크카드"),
            "bank_transfer": ("banktransfer", "transfer", "계좌이체", "이체"),
        }
        for standard, values in aliases.items():
            if compact in values:
                return standard
        if "체크카드" in compact or "checkcard" in compact or "debitcard" in compact:
            return "debit_card"
        if "카드승인" in compact or "카드결제" in compact or compact.endswith("카드") or compact.endswith("card"):
            return "credit_card"
    if field == "merchant":
        text = re.sub(r"(?:주식회사|\(?주\)?|㈜)", "", text)
        for descriptor in MERCHANT_IGNORED_DESCRIPTORS:
            text = text.replace(descriptor, "")
        compact = re.sub(r"[^0-9a-z가-힣]", "", text)
        return compact
    if field == "card_number":
        return re.sub(r"[^0-9*]", "", text)
    return text


def _merchant_box_state(expected_value: Any, pages: list[dict[str, Any]] | None) -> str:
    expected = re.sub(r"[^0-9a-z가-힣]", "", str(expected_value or "").lower())
    if not expected or not pages:
        return "unknown"
    box_texts = [
        str(item.get("text") or "")
        for page in pages if isinstance(page, dict)
        for item in (page.get("items") or []) if isinstance(item, dict)
    ]
    compact_boxes = [re.sub(r"[^0-9a-z가-힣]", "", text.lower()) for text in box_texts]
    if any(expected in box for box in compact_boxes):
        return "single"
    if expected in "".join(compact_boxes):
        return "split"
    tokens = [token for token in re.findall(r"[0-9A-Za-z가-힣]+", str(expected_value or "").lower()) if len(token) >= 2]
    token_boxes = [{index for index, box in enumerate(compact_boxes) if re.sub(r"[^0-9a-z가-힣]", "", token) in box} for token in tokens]
    if tokens and all(indexes for indexes in token_boxes) and len(set().union(*token_boxes)) > 1:
        return "split"
    return "unknown"


def _values_match(
    field: str,
    expected_value: Any,
    actual_value: Any,
    ocr_pages: list[dict[str, Any]] | None = None,
) -> bool:
    expected = _canonical(field, expected_value)
    actual = _canonical(field, actual_value)
    if expected == actual:
        return True
    if field in {"expense_category", "merchant", "name"} and normalization_equivalent(field, expected_value, actual_value):
        return True
    if field == "card_number" and expected and actual:
        expected_pattern = "".join("." if char == "*" else re.escape(char) for char in str(expected))
        return re.fullmatch(expected_pattern, str(actual)) is not None
    if field == "merchant" and expected and actual:
        if _merchant_box_state(expected_value, ocr_pages) != "split":
            return False
        if min(len(str(expected)), len(str(actual))) < 4:
            return False
        return expected in actual or actual in expected or SequenceMatcher(None, str(expected), str(actual)).ratio() >= 0.78
    if field != "name" or not expected or not actual:
        return False
    expected_tokens = _item_tokens(expected_value)
    actual_tokens = _item_tokens(actual_value)
    expected_numbers = {token for token in expected_tokens if re.fullmatch(r"[0-9]+(?:[.,][0-9]+)?", token)}
    actual_numbers = {token for token in actual_tokens if re.fullmatch(r"[0-9]+(?:[.,][0-9]+)?", token)}
    if expected_numbers and actual_numbers and expected_numbers != actual_numbers:
        return False
    expected_compact = re.sub(r"[^0-9a-z가-힣]", "", str(expected))
    actual_compact = re.sub(r"[^0-9a-z가-힣]", "", str(actual))
    if min(len(expected_compact), len(actual_compact)) >= 2 and (
        expected_compact in actual_compact or actual_compact in expected_compact
    ):
        return True
    if SequenceMatcher(None, expected_compact, actual_compact).ratio() >= 0.72:
        return True

    expected_words = expected_tokens - expected_numbers
    actual_words = actual_tokens - actual_numbers
    if min(len(expected_words), len(actual_words)) < 2:
        return False
    overlap = len(expected_words & actual_words)
    return overlap / min(len(expected_words), len(actual_words)) >= 0.75


def _match_items(expected_items: list[dict[str, Any]], predicted_items: list[dict[str, Any]]) -> dict[int, int]:
    candidates = []
    fields = ("name", "quantity", "unit_price", "total_amount")
    for expected_index, expected_item in enumerate(expected_items):
        for actual_index, actual_item in enumerate(predicted_items):
            score = sum(
                (3 if field == "name" else 1)
                for field in fields
                if field in expected_item and _values_match(field, expected_item.get(field), actual_item.get(field))
            )
            candidates.append((score, expected_index, actual_index))
    matches = {}
    used_actual = set()
    for _, expected_index, actual_index in sorted(candidates, reverse=True):
        if expected_index not in matches and actual_index not in used_actual:
            matches[expected_index] = actual_index
            used_actual.add(actual_index)
    return matches


def _empty(value: Any) -> bool:
    return value is None or value == "" or value == []


def _selection_rubric(
    prediction: dict[str, Any], truth: dict[str, Any], raw_prediction: dict[str, Any] | None,
    ocr_pages: list[dict[str, Any]] | None = None,
) -> dict[str, Any]:
    expected_items = truth.get("items") if isinstance(truth.get("items"), list) else []
    predicted_items = prediction.get("items") if isinstance(prediction.get("items"), list) else []
    item_matches = _match_items(expected_items, predicted_items)

    true_names = sum(
        _values_match("name", expected.get("name"), predicted_items[actual_index].get("name"))
        for index, expected in enumerate(expected_items)
        if (actual_index := item_matches.get(index)) is not None and "name" in expected
    )
    if not expected_items and not predicted_items:
        name_f1 = 1.0
    else:
        precision = true_names / len(predicted_items) if predicted_items else 0.0
        recall = true_names / len(expected_items) if expected_items else 0.0
        name_f1 = 2 * precision * recall / (precision + recall) if precision + recall else 0.0

    def item_accuracy(field: str) -> float:
        if not expected_items:
            return 1.0 if not predicted_items else 0.0
        correct = 0
        for index, expected in enumerate(expected_items):
            actual_index = item_matches.get(index)
            if actual_index is not None and _values_match(field, expected.get(field), predicted_items[actual_index].get(field)):
                correct += 1
        return correct / len(expected_items)

    def field_accuracy(field: str) -> float:
        expected = truth.get(field)
        actual = prediction.get(field)
        if field == "discount_amount" and _empty(expected):
            return 1.0 if _empty(actual) else 0.0
        return float(_values_match(field, expected, actual, ocr_pages))

    components = {
        "merchant": {"score": field_accuracy("merchant"), "weight": 5},
        "transaction_date": {"score": field_accuracy("transaction_date"), "weight": 5},
        "total_quantity": {"score": field_accuracy("total_quantity"), "weight": 4},
        "discount_amount": {"score": field_accuracy("discount_amount"), "weight": 4},
        "total_amount": {"score": field_accuracy("total_amount"), "weight": 10},
        "payment_method": {"score": field_accuracy("payment_method"), "weight": 3},
        "card_number": {"score": field_accuracy("card_number"), "weight": 4},
        "item_name_f1": {"score": name_f1, "weight": 12},
        "item_unit_price": {"score": item_accuracy("unit_price"), "weight": 8},
        "item_quantity": {"score": item_accuracy("quantity"), "weight": 7},
        "item_total_amount": {"score": item_accuracy("total_amount"), "weight": 8},
        "category": {"score": field_accuracy("expense_category"), "weight": 10},
    }

    raw = raw_prediction or prediction
    schema_checks = [
        "image" in raw or "source_filename" in raw,
        "merchant" in raw,
        "transaction_date" in raw,
        isinstance(raw.get("items"), list),
        "total_quantity" in raw or isinstance(raw.get("receipt_summary"), dict),
        "total_amount" in raw,
        "expense_category" in raw,
        "payment_method" in raw,
        "card_number" in raw,
    ]
    schema_rate = sum(schema_checks) / len(schema_checks)
    components["json_schema"] = {"score": schema_rate, "weight": 10}

    hallucinations = 0
    stability_targets = ("merchant", "discount_amount", "payment_method", "card_number")
    for field in stability_targets:
        if _empty(truth.get(field)) and not _empty(prediction.get(field)):
            hallucinations += 1
    if not expected_items and predicted_items:
        hallucinations += len(predicted_items)
    stability_rate = 1.0 if hallucinations == 0 else 0.0
    components["stability"] = {"score": stability_rate, "weight": 5}

    for component in components.values():
        component["points"] = component["score"] * component["weight"]
    extraction_score = sum(component["points"] for component in components.values())
    return {
        "version": "test01-test20-v1",
        "max_extraction_score": 95,
        "extraction_score": extraction_score,
        "extraction_rate": extraction_score / 95,
        "schema_rate": schema_rate,
        "total_amount_correct": bool(components["total_amount"]["score"]),
        "hallucination_count": hallucinations,
        "components": components,
    }


def score_fields(
    prediction: dict[str, Any], truth: dict[str, Any], raw_prediction: dict[str, Any] | None = None,
    ocr_pages: list[dict[str, Any]] | None = None,
) -> dict[str, Any]:
    details = {}
    evaluated = 0
    correct = 0
    for field in CORE_FIELDS:
        if field not in truth:
            continue
        matched = _values_match(field, truth.get(field), prediction.get(field), ocr_pages)
        details[field] = {"expected": truth.get(field), "actual": prediction.get(field), "correct": matched}
        evaluated += 1
        correct += int(matched)
    expected_items = truth.get("items")
    if isinstance(expected_items, list):
        predicted_items = [item for item in (prediction.get("items") or []) if isinstance(item, dict)] if isinstance(prediction.get("items"), list) else []
        item_fields = ("name", "quantity", "unit_price", "total_amount")
        item_details = []
        matches = _match_items(expected_items, predicted_items)
        count_matched = len(expected_items) == len(predicted_items)
        evaluated += 1
        correct += int(count_matched)
        for index, expected_item in enumerate(expected_items):
            actual_index = matches.get(index)
            actual_item = predicted_items[actual_index] if actual_index is not None else {}
            comparisons = {}
            for field in item_fields:
                if field not in expected_item:
                    continue
                matched = _values_match(field, expected_item.get(field), actual_item.get(field))
                comparisons[field] = {"expected": expected_item.get(field), "actual": actual_item.get(field), "correct": matched}
                evaluated += 1
                correct += int(matched)
            item_details.append({"index": index, "matched_actual_index": actual_index, "fields": comparisons})
        details["items"] = {
            "expected_count": len(expected_items),
            "actual_count": len(predicted_items),
            "count_correct": count_matched,
            "false_positive_count": max(0, len(predicted_items) - len(expected_items)),
            "items": item_details,
        }
    return {
        "correct_fields": correct,
        "evaluated_fields": evaluated,
        "field_accuracy": correct / evaluated if evaluated else 0,
        "complete_match": bool(evaluated) and correct == evaluated,
        "fields": details,
        "selection_rubric": _selection_rubric(prediction, truth, raw_prediction, ocr_pages),
    }


def _ocr_contains(text: str, field: str, value: Any) -> bool:
    if value is None or value == "":
        return False
    if field == "transaction_date":
        parts = re.findall(r"\d+", str(value))[:3]
        if len(parts) != 3:
            return False
        year, month, day = (int(part) for part in parts)
        date_pattern = (
            rf"(?<!\d){year:04d}\s*(?:년\s*|[-./]\s*)"
            rf"0?{month}\s*(?:월\s*|[-./]\s*)0?{day}\s*일?(?!\d)"
        )
        return re.search(date_pattern, text) is not None
    if field in NUMBER_FIELDS:
        expected_digits = re.sub(r"\D", "", str(value))
        return bool(expected_digits) and expected_digits in re.sub(r"\D", "", text)
    expected = re.sub(r"[^0-9a-z가-힣]", "", str(value).lower())
    source = re.sub(r"[^0-9a-z가-힣]", "", text.lower())
    return bool(expected) and expected in source


def estimate_ocr_impact(text: str, truth: dict[str, Any], score: dict[str, Any]) -> dict[str, Any]:
    """Estimate whether each extraction error likely began in OCR or the LLM.

    This is evidence-based attribution, not a CER/word-error-rate measurement.
    """
    fields = []

    def append(field: str, label: str, expected: Any, correct: bool) -> None:
        evidence_found = _ocr_contains(text, field, expected)
        if evidence_found and correct:
            status = "SUCCESS"
        elif evidence_found:
            status = "LIKELY_LLM_ERROR"
        elif correct:
            status = "LLM_RECOVERY"
        else:
            status = "LIKELY_OCR_ERROR"
        fields.append({
            "field": label,
            "expected": expected,
            "ocr_evidence_found": evidence_found,
            "llm_correct": bool(correct),
            "status": status,
        })

    score_fields_detail = score.get("fields") or {}
    for field, detail in score_fields_detail.items():
        if field == "items" or not isinstance(detail, dict):
            continue
        append(field, field, detail.get("expected"), bool(detail.get("correct")))

    items_detail = score_fields_detail.get("items") or {}
    for item in items_detail.get("items") or []:
        index = int(item.get("index") or 0)
        for field, detail in (item.get("fields") or {}).items():
            append(field, f"items[{index}].{field}", detail.get("expected"), bool(detail.get("correct")))

    counts = {status: sum(item["status"] == status for item in fields) for status in (
        "SUCCESS", "LIKELY_OCR_ERROR", "LIKELY_LLM_ERROR", "LLM_RECOVERY",
    )}
    evaluated = len(fields)
    return {
        "method": "FIELD_EVIDENCE_ESTIMATE",
        "notice": "OCR 원문에 정답 필드가 존재하는지 기반으로 한 원인 추정이며 CER/F1 점수가 아닙니다.",
        "evaluated_fields": evaluated,
        "ocr_evidence_rate": sum(item["ocr_evidence_found"] for item in fields) / evaluated if evaluated else 0,
        "counts": counts,
        "fields": fields,
    }


def verify_workbook(record: dict[str, Any]) -> dict[str, Any]:
    try:
        workbook = load_workbook(BytesIO(build_finance_workbook([record])), data_only=False)
        expected_sheet = SHEET_NAMES.get(record.get("document_type"))
        sheet = workbook[expected_sheet] if expected_sheet in workbook.sheetnames else workbook.active
        column_count = len(HEADERS_BY_TYPE.get(record.get("document_type"), [])) or 8
        headers = [sheet.cell(11, column).value for column in range(1, column_count + 1)]
        rows = [
            [sheet.cell(row, column).value for column in range(1, column_count + 1)]
            for row in range(12, max(12, sheet.max_row))
        ]
        sheet_previews = {}
        for worksheet in workbook.worksheets:
            if worksheet.title == SUMMARY_SHEET_NAME:
                preview_headers = [worksheet.cell(1, column).value for column in range(1, worksheet.max_column + 1)]
                preview_rows = [
                    [worksheet.cell(row, column).value for column in range(1, worksheet.max_column + 1)]
                    for row in range(2, worksheet.max_row + 1)
                ]
            else:
                preview_headers = [worksheet.cell(11, column).value for column in range(1, worksheet.max_column + 1)]
                preview_rows = [
                    [worksheet.cell(row, column).value for column in range(1, worksheet.max_column + 1)]
                    for row in range(12, max(12, worksheet.max_row))
                ]
            sheet_previews[worksheet.title] = {"headers": preview_headers, "rows": preview_rows}
        return {
            "success": expected_sheet in workbook.sheetnames and workbook.active.title == expected_sheet,
            "active_sheet": workbook.active.title,
            "expected_sheet": expected_sheet,
            "preview": {"headers": headers, "rows": rows},
            "sheet_previews": sheet_previews,
        }
    except Exception as exc:
        return {"success": False, "error": str(exc)}


async def evaluate_models(
    *, text: str, filename: str, truth: dict[str, Any], model_names: list[str],
    pages: list[dict[str, Any]] | None = None,
) -> list[dict[str, Any]]:
    truth = normalize_ground_truth(truth)
    results = []
    for model_name in model_names:
        logger.warning("Finance evaluation model start: model=%s filename=%s", model_name, filename)
        started = perf_counter()
        try:
            pure = await _classify_receipt_with_model(text, filename, model_name, pages)
            latency_ms = round((perf_counter() - started) * 1000)
            system = _normalize(dict(pure), filename, text)
            system_prediction = {field: system.get(field) for field in CORE_FIELDS}
            structured = system.get("structured_data") or {}
            system_prediction["items"] = structured.get("items") or []
            summary = structured.get("receipt_summary") if isinstance(structured.get("receipt_summary"), dict) else {}
            system_prediction["total_quantity"] = (
                structured.get("total_quantity")
                if structured.get("total_quantity") is not None
                else summary.get("stated_total_quantity")
            )
            system_prediction["discount_amount"] = structured.get("discount_amount")
            system_prediction["card_number"] = structured.get("card_number")
            system_score = score_fields(system_prediction, truth, pure, pages)
            structured_trace = structured.get("item_extraction_diagnostics") or {}
            pipeline_trace = {
                "llm": structured.get("llm_trace") or {},
                "validator": structured.get("validator_trace") or {},
                "deterministic_hints": structured.get("deterministic_hints") or {},
                "item_candidates": structured_trace.get("candidates") or [],
                "model_items": structured_trace.get("model_items") or [],
                "resolved_items": structured_trace.get("resolved_items") or [],
            }
            error_analysis = analyze_finance_evaluation_failure(
                ocr_text=text,
                ground_truth=truth,
                prediction=system_prediction,
                pipeline_trace=pipeline_trace,
            )
            results.append({
                "model_name": model_name,
                "success": True,
                "latency_ms": latency_ms,
                "system": {
                    "prediction": system_prediction,
                    "pipeline_trace": pipeline_trace,
                    "error_analysis": error_analysis,
                    "score": system_score,
                    "ocr_impact": estimate_ocr_impact(text, truth, system_score),
                    "workbook": verify_workbook(system),
                },
            })
        except Exception as exc:
            results.append({
                "model_name": model_name,
                "success": False,
                "latency_ms": round((perf_counter() - started) * 1000),
                "error": str(exc),
                "system": {"prediction": {}, "score": score_fields({}, truth), "workbook": {"success": False}},
            })
    return results
