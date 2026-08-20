from __future__ import annotations

from datetime import date
from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


SHEET_NAMES = {
    "EXPENSE_REPORT": "경비지출결의서",
    "TRAVEL_EXPENSE": "출장여비교통비정산서",
    "PURCHASE_REQUEST": "구매품의요청서",
    "WELFARE_BENEFIT": "복리후생비신청서",
}

TITLE_BY_TYPE = {
    "EXPENSE_REPORT": "경 비 지 출 결 의 서",
    "TRAVEL_EXPENSE": "출 장 / 여 비 교 통 비  정 산 서",
    "PURCHASE_REQUEST": "구 매 / 품 의 요 청 서",
    "WELFARE_BENEFIT": "복 리 후 생 비  신 청 서",
}

HEADERS_BY_TYPE = {
    "EXPENSE_REPORT": ["영수증 ID", "품목 순번", "결제일시", "상호명(가맹점)", "지출용도/품목명", "공급가액", "부가세", "합계금액", "증빙유형"],
    "TRAVEL_EXPENSE": ["영수증 ID", "품목 순번", "구분", "일자", "출발/도착지", "교통/숙박 수단", "상호(가맹점)", "금액", "증빙여부", "비고"],
    "PURCHASE_REQUEST": ["영수증 ID", "품목 순번", "품목명", "규격/옵션", "수량", "단위", "단가", "공급가액", "부가세", "합계금액", "비고"],
    "WELFARE_BENEFIT": ["영수증 ID", "품목 순번", "지원 항목(구분)", "결제일자", "내용(품목명/사유)", "결제처", "신청 금액", "증빙서류", "비고"],
}

SUMMARY_SHEET_NAME = "영수증요약"
SUMMARY_HEADERS = [
    "영수증 ID", "문서 유형", "거래일", "거래처",
    "OCR/LLM 판단 품목 수", "코드 계산 품목 행 수", "OCR/LLM 판단 총수량", "코드 계산 총수량", "단위 구성",
    "OCR/LLM 판단 총구매금액", "코드 계산 품목금액 합계", "차이금액", "검산 상태",
]


def _number(value: Any) -> float:
    try:
        return float(value or 0)
    except (TypeError, ValueError):
        return 0


def _record_rows(document_type: str, records: list[dict[str, Any]]) -> list[list[Any]]:
    rows: list[list[Any]] = []
    for record in records:
        data = record.get("structured_data") or {}
        items = data.get("items") if isinstance(data.get("items"), list) else []
        row_items = items or [None]
        receipt_id = str(record.get("document_id") or record.get("id") or "미확인")
        for item_index, item_value in enumerate(row_items, 1):
            item = item_value or {}
            is_single = len(row_items) == 1
            if document_type == "PURCHASE_REQUEST":
                quantity = _number(item.get("quantity")) or 1
                unit_price = _number(item.get("unit_price"))
                supply = _number(item.get("supply_amount")) or quantity * unit_price or (_number(record.get("supply_amount")) if is_single else 0)
                tax = _number(item.get("tax_amount")) or (_number(record.get("tax_amount")) if is_single else 0)
                total = _number(item.get("total_amount")) or supply + tax or (_number(record.get("total_amount")) if is_single else 0)
                rows.append([
                    receipt_id,
                    item_index,
                    item.get("name") or record.get("description"),
                    item.get("specification") or item.get("option"),
                    quantity,
                    item.get("unit") or "개",
                    unit_price,
                    supply,
                    tax,
                    total,
                    item.get("note"),
                ])
            elif document_type == "TRAVEL_EXPENSE":
                rows.append([
                    receipt_id, item_index, record.get("expense_category") or "교통비", record.get("transaction_date"),
                    data.get("route") or data.get("location"), data.get("transport_method") or data.get("service_type"),
                    record.get("merchant"), _number(item.get("total_amount")) or (_number(record.get("total_amount")) if is_single else 0),
                    data.get("evidence_status") or "첨부", item.get("name") or item.get("note") or data.get("note") or record.get("description"),
                ])
            elif document_type == "WELFARE_BENEFIT":
                rows.append([
                    receipt_id, item_index, record.get("expense_category") or "기타 복리후생", record.get("transaction_date"),
                    item.get("name") or record.get("description"), record.get("merchant"),
                    _number(item.get("total_amount")) or (_number(record.get("total_amount")) if is_single else 0),
                    data.get("evidence_type") or record.get("payment_method") or "영수증", item.get("note") or data.get("note"),
                ])
            else:
                quantity = _number(item.get("quantity")) or 1
                unit_price = _number(item.get("unit_price"))
                supply = _number(item.get("supply_amount")) or quantity * unit_price or (_number(record.get("supply_amount")) if is_single else 0)
                tax = _number(item.get("tax_amount")) or (_number(record.get("tax_amount")) if is_single else 0)
                total = _number(item.get("total_amount")) or supply + tax or (_number(record.get("total_amount")) if is_single else 0)
                rows.append([
                    receipt_id, item_index, record.get("transaction_date"), record.get("merchant"),
                    item.get("name") or record.get("description") or record.get("expense_category"), supply, tax, total,
                    record.get("payment_method") or "영수증",
                ])
    return rows


def _document_number(document_type: str, email: str, created_on: str) -> str:
    prefix = {"EXPENSE_REPORT": "EXP", "TRAVEL_EXPENSE": "TRV", "PURCHASE_REQUEST": "PUR", "WELFARE_BENEFIT": "WEL"}[document_type]
    checksum = sum((index + 1) * ord(character) for index, character in enumerate(email.lower())) % 1_000_000
    return f"{prefix}-{created_on.replace('-', '')}-{checksum:06d}"


def _record_period(records: list[dict[str, Any]]) -> str:
    dates = sorted(str(record.get("transaction_date")) for record in records if record.get("transaction_date"))
    if not dates:
        return "미확인"
    return dates[0] if len(dates) == 1 else f"{dates[0]} ~ {dates[-1]}"


def _summary_text(records: list[dict[str, Any]], field: str, fallback: str = "미입력") -> str:
    values = []
    for record in records:
        data = record.get("structured_data") or {}
        value = data.get(field)
        if value and str(value) not in values:
            values.append(str(value))
    return " / ".join(values[:3]) or fallback


def _optional_number(value: Any) -> float | None:
    if value is None or value == "":
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def _receipt_summary_rows(records: list[dict[str, Any]]) -> list[list[Any]]:
    rows = []
    for record in records:
        data = record.get("structured_data") or {}
        items = data.get("items") if isinstance(data.get("items"), list) else []
        stated = data.get("receipt_summary") if isinstance(data.get("receipt_summary"), dict) else {}
        stated_item_count = _optional_number(stated.get("stated_item_count"))
        stated_quantity = _optional_number(stated.get("stated_total_quantity"))
        receipt_total = _optional_number(stated.get("stated_total_amount"))
        if receipt_total is None:
            receipt_total = _number(record.get("total_amount"))

        extracted_quantity = sum(_number(item.get("quantity")) for item in items)
        extracted_total = sum(
            _number(item.get("total_amount"))
            or _number(item.get("supply_amount")) + _number(item.get("tax_amount"))
            or _number(item.get("quantity")) * _number(item.get("unit_price"))
            for item in items
        )
        units: dict[str, float] = {}
        for item in items:
            unit = str(item.get("unit") or "단위 미확인")
            units[unit] = units.get(unit, 0) + _number(item.get("quantity"))
        unit_summary = " / ".join(
            f"{unit} {quantity:g}" for unit, quantity in units.items()
        ) or None

        mismatches = []
        if stated_item_count is not None and stated_item_count != len(items):
            mismatches.append("품목 수")
        if stated_quantity is not None and stated_quantity != extracted_quantity:
            mismatches.append("총수량")
        if receipt_total and abs(receipt_total - extracted_total) > 0.01:
            mismatches.append("금액")
        if not items:
            mismatches.append("품목 없음")
        status = "일치" if not mismatches else f"확인 필요: {', '.join(mismatches)}"
        rows.append([
            str(record.get("document_id") or record.get("id") or "미확인"),
            SHEET_NAMES.get(record.get("document_type"), record.get("document_type")),
            record.get("transaction_date"), record.get("merchant"),
            stated_item_count, len(items), stated_quantity, extracted_quantity, unit_summary,
            receipt_total, extracted_total, receipt_total - extracted_total, status,
        ])
    return rows


def _style_summary_sheet(ws, records: list[dict[str, Any]]) -> None:
    dark = "1F4E78"
    line = Side(style="thin", color="AAB7C4")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A2"
    for column, label in enumerate(SUMMARY_HEADERS, 1):
        cell = ws.cell(1, column, label)
        cell.fill = PatternFill("solid", fgColor=dark)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(top=line, bottom=line, left=line, right=line)
    for row_index, values in enumerate(_receipt_summary_rows(records), 2):
        for column, value in enumerate(values, 1):
            cell = ws.cell(row_index, column, value)
            cell.border = Border(top=line, bottom=line, left=line, right=line)
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            if column in {5, 6, 7, 8, 10, 11, 12}:
                cell.number_format = '#,##0.###'
        status_cell = ws.cell(row_index, 13)
        status_cell.font = Font(bold=True, color="14835B" if values[12] == "일치" else "C9474F")
    widths = [38, 22, 16, 26, 14, 16, 14, 14, 24, 18, 18, 16, 30]
    for index, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(index)].width = width
    ws.auto_filter.ref = f"A1:M{max(len(records) + 1, 2)}"


def _style_sheet(ws, document_type: str, records: list[dict[str, Any]], author: dict[str, str]) -> None:
    dark = "1F4E78"
    light = "D9EAF7"
    line = Side(style="thin", color="AAB7C4")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A12"
    ws.merge_cells("A2:D3")
    ws["A2"] = TITLE_BY_TYPE[document_type]
    ws["A2"].font = Font(name="맑은 고딕", size=18, bold=True, color="FFFFFF")
    ws["A2"].fill = PatternFill("solid", fgColor=dark)
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells("E2:E4")
    ws["E2"] = "결\n\n재"
    ws["E2"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for column, label in zip(("F", "G", "H"), ("기안자", "검토자", "승인자")):
        ws.merge_cells(f"{column}3:{column}4")
        ws[f"{column}2"] = label
        ws[f"{column}2"].font = Font(bold=True)
        ws[f"{column}2"].alignment = Alignment(horizontal="center")
    created_on = date.today().isoformat()
    author_name = author.get("name") or author.get("email") or "작성자 미확인"
    author_email = author.get("email") or "미등록"
    department = author.get("department") or "미등록"
    ws["F3"] = author_name

    ws["A5"], ws["B5"], ws["E5"], ws["F5"] = "문서번호", _document_number(document_type, author_email, created_on), "처리상태", "확정"
    ws.merge_cells("B5:D5"); ws.merge_cells("F5:H5")

    if document_type == "EXPENSE_REPORT":
        metadata = [("기안일자", created_on, "부서명", department), ("기안자", author_name, "작성자 이메일", author_email), ("정산기간", _record_period(records), "영수증 수", f"{len(records)}건")]
    elif document_type == "TRAVEL_EXPENSE":
        metadata = [("출장자", author_name, "소속부서", department), ("출장목적", _summary_text(records, "note", "영수증 자동 정산"), "작성자 이메일", author_email), ("출장기간", _record_period(records), "출장지", _summary_text(records, "location"))]
    elif document_type == "PURCHASE_REQUEST":
        merchants = " / ".join(dict.fromkeys(str(record.get("merchant")) for record in records if record.get("merchant"))) or "미확인"
        receipt_total = sum(_number(record.get("total_amount")) for record in records)
        metadata = [("거래일", _record_period(records), "거래처", merchants), ("작성부서", department, "작성자", author_name), ("영수증 총액", receipt_total, "비용구분", records[0].get("expense_category", "미입력") if records else "미입력")]
    else:
        metadata = [("신청일자", created_on, "소속부서", department), ("신청인", author_name, "작성자 이메일", author_email), ("신청기간", _record_period(records), "신청건수", f"{len(records)}건")]

    for row_number, (left_label, left_value, right_label, right_value) in enumerate(metadata, 6):
        ws[f"A{row_number}"], ws[f"B{row_number}"], ws[f"E{row_number}"], ws[f"F{row_number}"] = left_label, left_value, right_label, right_value
        ws.merge_cells(f"B{row_number}:D{row_number}"); ws.merge_cells(f"F{row_number}:H{row_number}")
    for row_number in range(5, 9):
        for cell in ws[row_number]:
            cell.fill = PatternFill("solid", fgColor=light if row_number > 5 else "EAF1F7")
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws[f"A{row_number}"].font = Font(bold=True); ws[f"E{row_number}"].font = Font(bold=True)

    header_row = 11
    for column, label in enumerate(HEADERS_BY_TYPE[document_type], 1):
        cell = ws.cell(header_row, column, label)
        cell.fill = PatternFill("solid", fgColor=dark)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(top=line, bottom=line, left=line, right=line)

    rows = _record_rows(document_type, records)
    column_count = len(HEADERS_BY_TYPE[document_type])
    if not rows:
        rows = [[None] * column_count]
    for row_index, values in enumerate(rows, header_row + 1):
        for column, value in enumerate(values, 1):
            cell = ws.cell(row_index, column, value)
            cell.border = Border(top=line, bottom=line, left=line, right=line)
            money_columns = {
                "EXPENSE_REPORT": {6, 7, 8},
                "TRAVEL_EXPENSE": {8},
                "PURCHASE_REQUEST": {7, 8, 9, 10},
                "WELFARE_BENEFIT": {7},
            }[document_type]
            cell.alignment = Alignment(horizontal="right" if column in money_columns else "left", vertical="center", wrap_text=True)
            if column in money_columns:
                cell.number_format = '#,##0'

    total_row = header_row + len(rows) + 1
    total_label_end = {"EXPENSE_REPORT": 5, "TRAVEL_EXPENSE": 7, "PURCHASE_REQUEST": 7, "WELFARE_BENEFIT": 6}[document_type]
    ws.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=total_label_end)
    ws.cell(total_row, 1, "합 계")
    ws.cell(total_row, 1).font = Font(bold=True)
    ws.cell(total_row, 1).fill = PatternFill("solid", fgColor=light)
    if document_type == "PURCHASE_REQUEST":
        total_columns = (8, 9, 10)
        for column in total_columns:
            letter = get_column_letter(column)
            ws.cell(total_row, column, f"=SUM({letter}{header_row + 1}:{letter}{total_row - 1})")
            ws.cell(total_row, column).number_format = '#,##0" 원"'
    elif document_type == "EXPENSE_REPORT":
        for column in (6, 7, 8):
            letter = get_column_letter(column)
            ws.cell(total_row, column, f"=SUM({letter}{header_row + 1}:{letter}{total_row - 1})")
            ws.cell(total_row, column).number_format = '#,##0" 원"'
    else:
        amount_column = 8 if document_type == "TRAVEL_EXPENSE" else 7
        letter = get_column_letter(amount_column)
        ws.cell(total_row, amount_column, f"=SUM({letter}{header_row + 1}:{letter}{total_row - 1})")
        ws.cell(total_row, amount_column).number_format = '#,##0" 원"'
    for column in range(1, column_count + 1):
        ws.cell(total_row, column).border = Border(top=Side(style="medium", color=dark))

    widths = {
        "EXPENSE_REPORT": [38, 10, 16, 24, 30, 18, 16, 18, 18],
        "TRAVEL_EXPENSE": [38, 10, 14, 16, 24, 22, 24, 18, 16, 28],
        "PURCHASE_REQUEST": [38, 10, 28, 24, 10, 10, 16, 18, 16, 18, 28],
        "WELFARE_BENEFIT": [38, 10, 20, 16, 30, 24, 18, 18, 24],
    }[document_type]
    for index, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(index)].width = width
    ws.row_dimensions[2].height = 30
    ws.row_dimensions[11].height = 34
    ws.auto_filter.ref = f"A11:{get_column_letter(column_count)}{max(total_row - 1, 12)}"


def build_finance_workbook(records: list[dict[str, Any]], author: dict[str, str] | None = None) -> bytes:
    workbook = Workbook()
    workbook.remove(workbook.active)
    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True
    workbook.calculation.calcMode = "auto"
    for document_type, sheet_name in SHEET_NAMES.items():
        ws = workbook.create_sheet(sheet_name)
        matching = [record for record in records if record.get("document_type") == document_type]
        _style_sheet(ws, document_type, matching, author or {})
    summary_sheet = workbook.create_sheet(SUMMARY_SHEET_NAME)
    _style_summary_sheet(summary_sheet, records)
    first_document_type = records[0].get("document_type") if records else None
    if first_document_type in SHEET_NAMES:
        workbook.active = list(SHEET_NAMES).index(first_document_type)
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()
