from io import BytesIO
from pathlib import Path
import sys
import unittest

from openpyxl import load_workbook

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from app.services.finance_workbook_service import build_finance_workbook  # noqa: E402


class FinanceWorkbookServiceTests(unittest.TestCase):
    def test_builds_four_standard_sheets_and_expense_totals(self):
        content = build_finance_workbook([
            {
                "document_id": "receipt-expense-001",
                "document_type": "EXPENSE_REPORT",
                "expense_category": "회의비",
                "merchant": "테스트카페",
                "transaction_date": "2026-08-18",
                "supply_amount": 10000,
                "tax_amount": 1000,
                "total_amount": 11000,
                "payment_method": "법인카드",
                "description": "고객 미팅",
                "structured_data": {},
                "created_at": "2026-08-18T10:00:00+09:00",
            }
        ], author={"name": "테스트 사용자", "email": "tester@example.com"})
        workbook = load_workbook(BytesIO(content), data_only=False)
        self.assertEqual(
            workbook.sheetnames,
            ["경비지출결의서", "출장여비교통비정산서", "구매품의요청서", "복리후생비신청서", "영수증요약"],
        )
        sheet = workbook["경비지출결의서"]
        self.assertEqual(sheet["A5"].value, "문서번호")
        self.assertTrue(sheet["B5"].value.startswith("EXP-"))
        self.assertEqual(sheet["B7"].value, "테스트 사용자")
        self.assertEqual(sheet["F7"].value, "tester@example.com")
        self.assertEqual(sheet["F3"].value, "테스트 사용자")
        self.assertEqual([sheet.cell(11, column).value for column in range(1, 10)], [
            "영수증 ID", "품목 순번", "결제일시", "상호명(가맹점)", "지출용도/품목명", "공급가액", "부가세", "합계금액", "증빙유형",
        ])
        self.assertEqual(sheet["A12"].value, "receipt-expense-001")
        self.assertEqual(sheet["B12"].value, 1)
        self.assertEqual(sheet["D12"].value, "테스트카페")
        self.assertEqual(sheet["H12"].value, 11000)
        self.assertEqual(sheet["H13"].value, "=SUM(H12:H12)")

    def test_places_confirmed_travel_meal_in_travel_sheet_columns(self):
        content = build_finance_workbook([{
            "document_id": "receipt-travel-001",
            "document_type": "TRAVEL_EXPENSE",
            "expense_category": "일비/식대",
            "merchant": "카페마마스 광화문점",
            "transaction_date": "2025-10-05",
            "total_amount": 31400,
            "description": "서울 출장 식비",
            "structured_data": {"location": "서울 종로구", "evidence_status": "첨부"},
            "created_at": "2025-10-05T16:50:00+09:00",
        }])
        workbook = load_workbook(BytesIO(content), data_only=False)
        sheet = workbook["출장여비교통비정산서"]
        self.assertEqual(workbook.active.title, "출장여비교통비정산서")
        self.assertEqual(sheet["A12"].value, "receipt-travel-001")
        self.assertEqual(sheet["B12"].value, 1)
        self.assertEqual(sheet["C12"].value, "일비/식대")
        self.assertEqual(sheet["D12"].value, "2025-10-05")
        self.assertEqual(sheet["E12"].value, "서울 종로구")
        self.assertEqual(sheet["G12"].value, "카페마마스 광화문점")
        self.assertEqual(sheet["H12"].value, 31400)
        self.assertEqual(sheet["J12"].value, "서울 출장 식비")

    def test_writes_every_purchase_item_to_its_own_row(self):
        content = build_finance_workbook([{
            "document_id": "receipt-purchase-001",
            "document_type": "PURCHASE_REQUEST",
            "expense_category": "사무용품",
            "merchant": "한국문구",
            "transaction_date": "2026-08-20",
            "total_amount": 8800,
            "structured_data": {"items": [
                {"name": "볼펜", "specification": "검정 0.5mm", "quantity": 2, "unit": "자루", "unit_price": 1100, "supply_amount": 2000, "tax_amount": 200, "total_amount": 2200},
                {"name": "노트", "specification": "A5 80매", "quantity": 3, "unit": "권", "unit_price": 2200, "supply_amount": 6000, "tax_amount": 600, "total_amount": 6600},
            ]},
        }])
        workbook = load_workbook(BytesIO(content), data_only=False)
        sheet = workbook["구매품의요청서"]

        self.assertEqual([sheet.cell(11, column).value for column in range(1, 12)], [
            "영수증 ID", "품목 순번", "품목명", "규격/옵션", "수량", "단위", "단가", "공급가액", "부가세", "합계금액", "비고",
        ])
        self.assertEqual([sheet["A12"].value, sheet["B12"].value, sheet["C12"].value, sheet["D12"].value, sheet["E12"].value, sheet["F12"].value], ["receipt-purchase-001", 1, "볼펜", "검정 0.5mm", 2, "자루"])
        self.assertEqual([sheet["A13"].value, sheet["B13"].value, sheet["C13"].value, sheet["D13"].value, sheet["E13"].value, sheet["F13"].value], ["receipt-purchase-001", 2, "노트", "A5 80매", 3, "권"])
        self.assertEqual(sheet["H14"].value, "=SUM(H12:H13)")
        self.assertEqual(sheet["I14"].value, "=SUM(I12:I13)")
        self.assertEqual(sheet["J14"].value, "=SUM(J12:J13)")
        self.assertEqual(sheet["B8"].value, 8800)

    def test_uses_receipt_id_and_item_sequence_for_all_document_types(self):
        records = []
        for document_type in ("EXPENSE_REPORT", "TRAVEL_EXPENSE", "WELFARE_BENEFIT"):
            records.append({
                "document_id": f"receipt-{document_type.lower()}",
                "document_type": document_type,
                "expense_category": "테스트",
                "merchant": "테스트 거래처",
                "transaction_date": "2026-08-20",
                "total_amount": 3000,
                "structured_data": {"items": [
                    {"name": "품목 A", "total_amount": 1000},
                    {"name": "품목 B", "total_amount": 2000},
                ]},
            })
        workbook = load_workbook(BytesIO(build_finance_workbook(records)), data_only=False)

        for document_type, sheet_name in (
            ("EXPENSE_REPORT", "경비지출결의서"),
            ("TRAVEL_EXPENSE", "출장여비교통비정산서"),
            ("WELFARE_BENEFIT", "복리후생비신청서"),
        ):
            sheet = workbook[sheet_name]
            receipt_id = f"receipt-{document_type.lower()}"
            self.assertEqual([sheet["A12"].value, sheet["B12"].value], [receipt_id, 1])
            self.assertEqual([sheet["A13"].value, sheet["B13"].value], [receipt_id, 2])

    def test_adds_receipt_summary_with_stated_and_extracted_reconciliation(self):
        record = {
            "document_id": "receipt-summary-001",
            "document_type": "PURCHASE_REQUEST",
            "merchant": "테스트 상점",
            "transaction_date": "2026-08-20",
            "total_amount": 81300,
            "structured_data": {
                "receipt_summary": {
                    "stated_item_count": 2,
                    "stated_total_quantity": 7,
                    "stated_total_amount": 81300,
                },
                "items": [
                    {"name": "품목 A", "quantity": 2, "unit": "개", "total_amount": 33600},
                    {"name": "품목 B", "quantity": 1, "unit": "개", "total_amount": 11000},
                    {"name": "품목 C", "quantity": 1, "unit": "개", "total_amount": 8500},
                    {"name": "품목 D", "quantity": 1, "unit": "개", "total_amount": 14000},
                    {"name": "품목 E", "quantity": 1, "unit": "개", "total_amount": 49000},
                ],
            },
        }
        workbook = load_workbook(BytesIO(build_finance_workbook([record])), data_only=False)
        sheet = workbook["영수증요약"]

        self.assertEqual([sheet.cell(1, column).value for column in range(1, 14)], [
            "영수증 ID", "문서 유형", "거래일", "거래처", "OCR/LLM 판단 품목 수", "코드 계산 품목 행 수",
            "OCR/LLM 판단 총수량", "코드 계산 총수량", "단위 구성", "OCR/LLM 판단 총구매금액", "코드 계산 품목금액 합계", "차이금액", "검산 상태",
        ])
        self.assertEqual(sheet["A2"].value, "receipt-summary-001")
        self.assertEqual([sheet["E2"].value, sheet["F2"].value], [2, 5])
        self.assertEqual([sheet["G2"].value, sheet["H2"].value], [7, 6])
        self.assertEqual(sheet["J2"].value, 81300)
        self.assertIn("품목 수", sheet["M2"].value)
        self.assertIn("총수량", sheet["M2"].value)
        self.assertIn("금액", sheet["M2"].value)


if __name__ == "__main__":
    unittest.main()
