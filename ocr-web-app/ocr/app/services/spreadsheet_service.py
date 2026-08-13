from datetime import date, datetime, time
from pathlib import Path

from openpyxl import load_workbook

from app.schemas.ocr import OCRItem, OCRPage


MAX_ROWS_PER_SHEET = 10_000
MAX_COLUMNS_PER_SHEET = 200
MAX_CELLS_PER_WORKBOOK = 200_000


def _display_value(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, (datetime, date, time)):
        return value.isoformat(sep=" ") if isinstance(value, datetime) else value.isoformat()
    return str(value)


def extract_spreadsheet(file_path: Path) -> list[OCRPage]:
    """Extract worksheet values and cell coordinates without executing macros."""
    workbook = load_workbook(
        filename=file_path,
        read_only=True,
        data_only=False,
        keep_links=False,
    )
    pages: list[OCRPage] = []
    processed_cells = 0

    try:
        for sheet_number, worksheet in enumerate(workbook.worksheets, start=1):
            rows: list[list[str]] = []
            items: list[OCRItem] = []

            for row_number, cells in enumerate(
                worksheet.iter_rows(max_col=MAX_COLUMNS_PER_SHEET),
                start=1,
            ):
                if row_number > MAX_ROWS_PER_SHEET:
                    break

                values = [_display_value(cell.value) for cell in cells]
                while values and not values[-1]:
                    values.pop()

                rows.append(values)
                for column_number, value in enumerate(values, start=1):
                    source_cell = cells[column_number - 1]
                    processed_cells += 1
                    if processed_cells > MAX_CELLS_PER_WORKBOOK:
                        raise ValueError("Excel 파일에서 처리 가능한 셀 수를 초과했습니다.")
                    if not value:
                        continue
                    items.append(
                        OCRItem(
                            text=value,
                            confidence=1.0,
                            bbox=[
                                [column_number - 1, row_number - 1],
                                [column_number, row_number - 1],
                                [column_number, row_number],
                                [column_number - 1, row_number],
                            ],
                            cell=source_cell.coordinate,
                            row=row_number,
                            column=column_number,
                        )
                    )

                if not values and rows:
                    # Keep interior empty rows, then remove trailing empty rows below.
                    continue

            while rows and not rows[-1]:
                rows.pop()

            text = "\n".join(" | ".join(row) for row in rows if any(row))
            pages.append(
                OCRPage(
                    page=sheet_number,
                    text=text,
                    items=items,
                    sheet_name=worksheet.title,
                    rows=rows,
                )
            )
    finally:
        workbook.close()

    return pages
